# coding:utf-8
import shutil
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import requests
from requests import Response
import os
import sys
from pathlib import Path

def process(col_letter, data_row_start, fwcli_template):
    print(f"processing: col_letter {col_letter} data_row_start {data_row_start} fwcli_template {fwcli_template}")

    if fwcli_template != "None" and fwcli_template != "":
        raise Exception("fwcli_template存在时的逻辑未处理!")
    
        

    pass

def checkpath(path_src):
    save = False
    workbook_src = load_workbook(path_src, data_only=True)
    for shert_name in workbook_src.sheetnames:
        sheet_src = workbook_src[shert_name]
        start_row = 0
        template_row = 0
        
        for row in range(2, sheet_src.max_row + 1):
            cell_name = "A" + str(row) # 取第一行
            cell_value = str(sheet_src[cell_name].value)
            if "fwcli" not in cell_value:
                start_row = row
                break
            else:
                if "fwcli_template" in cell_value:
                    template_row = row

        for col in range(1, sheet_src.max_column + 1):
            letter = get_column_letter(col)
            cell_name = letter + "1" # 取第一列
            cell_value = str(sheet_src[cell_name].value)
            if "MAP_Expression" in cell_value:
                print(f"{path_src} {cell_value} start_row:{start_row} template_row:{template_row}")
                fwcli_template = "" if template_row == 0 else str(sheet_src[letter + str(template_row)].value)
                process(letter, start_row, fwcli_template)
                save = True
            # sheet_src[cell_name].value = "X_MAP_Expression"
    if save:
        workbook_src.save(path_src) # 储存
    else:
        os.remove(path_src)

if __name__ == "__main__":
    path_src = sys.argv[1]
    print(path_src)

    t = path_src.split("\\")
    t[-1] = "xlsx_cexpression"
    path_src = "\\".join(t)
    
    print(path_src)
    if os.path.isdir(path_src):
        shutil.rmtree(path_src, ignore_errors=True)
    shutil.copytree(sys.argv[1], path_src)    

    if os.path.isdir(path_src):
        files = os.listdir(path_src)
        for i,filename in enumerate(files):
            # print("checkpath {}, {} / {}".format(filename, i+1, len(files)))
            checkpath(os.path.join(path_src, filename))
    else:
        checkpath(path_src)
    os.chdir(Path(path_src).parent)
    os.system("fwcli xlsxtotsvall -s xlsx_cexpression\ -d config")
    
    print("done!")

