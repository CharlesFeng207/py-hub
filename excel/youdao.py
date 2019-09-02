# -*- coding: utf-8 -*-
import sys
import uuid
import requests
import hashlib
import time
import json
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

YOUDAO_URL = 'http://openapi.youdao.com/api'
APP_KEY = ''
APP_SECRET = ''


def encrypt(signStr):
    hash_algorithm = hashlib.sha256()
    hash_algorithm.update(signStr.encode('utf-8'))
    return hash_algorithm.hexdigest()


def truncate(q):
    if q is None:
        return None
    size = len(q)
    return q if size <= 20 else q[0:10] + str(size) + q[size - 10:size]


def do_request(data):
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    return requests.post(YOUDAO_URL, data=data, headers=headers)


def translate(q, to):
    data = {}
    data['from'] = 'zh-CHS'
    data['to'] = to
    data['signType'] = 'v3'
    curtime = str(int(time.time()))
    data['curtime'] = curtime
    salt = str(uuid.uuid1())
    signStr = APP_KEY + truncate(q) + salt + curtime + APP_SECRET
    sign = encrypt(signStr)
    data['appKey'] = APP_KEY
    data['q'] = q
    data['salt'] = salt
    data['sign'] = sign

    print("send {} to {}".format(q, to))

    response = do_request(data)
    if response.status_code == 200:
        obj = json.loads(response.text)
        if obj["errorCode"] == "0":
            result = obj["translation"][0]
            print(result)
            return result

    return ""


def process_cell(cn, cell, lan):
    if cell.value is None or cell.value == "":
        cell.value = translate(cn, lan)


if __name__ == "__main__":

    # print(translate("发动计谋，对敌方造成<color=#189DE2>{0}</color>伤害，同时随机使两名友军双攻各增加4%，并使自身受到伤害减少10%，持续3秒", "en"))
    # print(translate("凑齐<color=#67d940>20个碎片</color>可招募将领<color=#d85fee>逢纪</color>，分解该碎片可获得<color=#67d940>20个将魂</color>", "en"))
    # print(translate("抡起双戟，对当前攻击目标造成<color=#189DE2>{0}</color>伤害，同时回复自身30%血量", "en"))
    path_src = "Language.xlsx"
    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    target_letter = "B"
    start_row = 3

    for row in range(start_row, sheet_src.max_row + 1):
        cell = sheet_src["{}{}".format(target_letter, row)]
        if cell.value is None:
            continue
        
        process_cell(str(cell.value), sheet_src["{}{}".format("C", row)], "en")
        process_cell(str(cell.value), sheet_src["{}{}".format("E", row)], "ja")
        process_cell(str(cell.value), sheet_src["{}{}".format("F", row)], "ko")

    workbook_src.save(path_src)

    print("done!")