# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import requests
from requests import Response
import os
import sys


def process(path_src):
    name_src = os.path.basename(path_src)

    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(
        filter(lambda x: x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    target_cols = []
    i = 1
    while True:
        cell = sheet_src["{}{}".format(get_column_letter(i), 2)]
        if cell.value is None:
            break
        if cell.comment is not None and "language done" in cell.comment.content:
            target_cols.append(get_column_letter(i+1))
        i += 1

    print(target_cols)

    if len(target_cols) > 0 and sheet_src["A1"].value != "#comment":
        sheet_src.insert_cols(0, 1)
        sheet_src["A1"].value = "#comment"

    for target_letter in target_cols:
        for row in range(5, sheet_src.max_row + 1):
            cell_name = "{}{}".format(target_letter, row)
            cell = sheet_src[cell_name]
           
            print("{} cur:{} total:{} value:{} cell:{}".format(
                name_src, row, sheet_src.max_row, cell.value, cell_name))

            cell_value = int(cell.value) if cell.value is not None else 0 

            comment_cell = sheet_src["A{}".format(row)]
            if comment_cell.value is None:
                comment_cell.value = ""

            if cell_value != 0:
                respond = None
                while True:
                    # url = 'http://139.155.88.114:5000/query_cn'
                    url = 'http://127.0.0.1:5000/query_cn'
                    respond = requests.get(url, {"lanId": cell_value})
                    if respond.status_code == 200:
                        break
                comment_cell.value += (" " + respond.text)
                print(comment_cell.value)
                pass
            
            pass
        pass

    workbook_src.save(name_src)
    print("saved")


if __name__ == "__main__":
    path_src = sys.argv[1]
    print(path_src)

    if os.path.isdir(path_src):
        files = os.listdir(path_src)
        for i, filename in enumerate(files):
            print("process {}, {} / {}".format(filename, i+1, len(files)))
            process(os.path.join(path_src, filename))
    else:
        process(path_src)
        pass

    print("done!")
