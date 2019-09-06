# -*- coding: utf-8 -*-
import sys
import uuid
import requests
import hashlib
import time
import json
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os
from openpyxl.styles import PatternFill, colors


if __name__ == "__main__":
    outputFolder = "done"
    files = os.listdir(outputFolder)
    pending = {}  # id -> {} (cn en bgcolor)

    total = 0
    markcount = 0
    for i, filename in enumerate(files):
        print("process {}, {} / {}".format(filename, i+1, len(files)))
        workbook_p = load_workbook(os.path.join(
            outputFolder, filename), data_only=True)
        s = workbook_p["work"]
        for row in range(1, s.max_row + 1):
            id_cell = s["A{}".format(row)]
            if id_cell is None or id_cell.value is None:
                continue
            cell_id = int(id_cell.value)
            cn_cell = s["B{}".format(row)]
            en_cell = s["C{}".format(row)]

            mark = False

            if id_cell.fill.end_color.index != colors.BLACK:
                mark = True

            if cn_cell.fill.end_color.index != colors.BLACK:
                mark = True

            if en_cell.fill.end_color.index != colors.BLACK:
                mark = True

            obj = {"cn": cn_cell.value, "en": en_cell.value, "mark": mark}
            pending[cell_id] = obj
            print(f"id:{cell_id} obj:{obj}")

            total += 1
            if mark:
                markcount += 1
            pass
        pass

    print(markcount, total)

    if input("press y to contionue:") == "y":
        path_src = "Language.xlsx"
        workbook_src = load_workbook(path_src, data_only=True)
        shert_name = list(
            filter(lambda x: x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]

        sheet_src = workbook_src[shert_name]
        start_row = 3

        marked = []
        for row in range(start_row, sheet_src.max_row + 1):
            i = int(sheet_src["A{}".format(row)].value)

            if i in pending:
                oldcn = sheet_src["B{}".format(row)].value
                
                if oldcn != pending[i]["cn"]:
                    sheet_src["B{}".format(row)].value = pending[i]["cn"]
                    print(f"{oldcn} -> {pending[i]['cn']}")

                olden = sheet_src["C{}".format(row)].value
                if olden != pending[i]["en"]:
                    sheet_src["C{}".format(row)].value = pending[i]["en"]
                    print(f"{olden} -> {pending[i]['en']}")

                if pending[i]["mark"]:
                    myFill = PatternFill(
                        start_color='F5DEB300', end_color='F5DEB300', fill_type='solid')
                    sheet_src["C{}".format(row)].fill = myFill
                    marked.append(row)

        workbook_src.save(path_src)

        print("marked count {}!".format(len(marked)))

        if input("press y to contionue:") == "y":
            nwb = Workbook()
            nsheet = nwb.active

            i = 1
            for r in marked:
                nsheet["A{}".format(
                    i)].value = sheet_src["A{}".format(r)].value
                nsheet["B{}".format(
                    i)].value = sheet_src["B{}".format(r)].value
                # nsheet["C{}".format(i)].value = sheet_src["C{}".format(r)].value
                i += 1

            nwb.save("translate.xlsx")

    print("done!")
