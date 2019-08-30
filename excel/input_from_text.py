# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os
import sys

if __name__ == "__main__":
    path_src = "input.txt"
    target_letter = sys.argv[1]
    print(target_letter)

    with open(path_src, "r", encoding='utf-8') as f:
        s = f.read()
        outList = s.split("@")
        final_list = list(map(lambda x : x.strip(), outList))
        print(len(final_list))

    if input("press y to contionue:") == "y":
        start_row = 4
        target_excel = "Language.xlsx"
        workbook_src = load_workbook(target_excel, data_only=True)
        shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
        print(shert_name)

        sheet_src = workbook_src[shert_name]

        for i, c in enumerate(final_list):
            row = i + start_row
            if row <= sheet_src.max_row:
                cell = sheet_src["{}{}".format(target_letter, row)]
                cell.value = c
                print("row:{} -> {}".format(row, c))
            else:
                print("row:{} -> None".format(row))
        
        workbook_src.save(target_excel)

    print("done!")




