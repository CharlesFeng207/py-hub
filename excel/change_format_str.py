# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import requests
from requests import Response
import os
import sys


def process(path_src):

    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    target_letter = "C"
    start_row = 5

    for row in range(start_row, sheet_src.max_row + 1):
        cell = sheet_src["{}{}".format(target_letter, row)]
        if cell.value is None:
            continue

        cell_value = str(cell.value)
        result = cell_value[0:cell_value.index(">") + 1] + "{0}" + cell_value[cell_value.rindex("<"):len(cell_value)]
        print(result)
        cell.value = result

    workbook_src.save(path_src)

if __name__ == "__main__":
    path_src = sys.argv[1]
    print(path_src)

    if os.path.isdir(path_src):
        files = os.listdir(path_src)
        for i,filename in enumerate(files):
            print("process {}, {} / {}".format(filename, i+1, len(files)))
            process(os.path.join(path_src, filename))
    else:
        process(path_src)
    
    print("done!")

