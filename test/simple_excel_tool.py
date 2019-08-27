# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import requests
from requests import Response
import os
import sys


def process(path_src):

    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    target_letter = "A"
    output_letter = "B"

    for row in range(1, sheet_src.max_row + 1):
        cell = sheet_src["{}{}".format(target_letter, row)]
        print("{} cur:{} total:{} value:{}".format(path_src, row, sheet_src.max_row, cell.value))
        if cell.value is not None:
            respond = None
            while True:
                isLast = "1" if row == sheet_src.max_row else "0"
                url = 'http://139.155.88.114:5000/query'
                respond = requests.get(url, {"content":cell.value, "save":isLast})
                if respond.status_code == 200:
                    break
            sheet_src["{}{}".format(output_letter, row)].value = respond.text
        else:
            sheet_src["{}{}".format(output_letter, row)].value = 0

    workbook_src.save(path_src)

if __name__ == "__main__":
    path_src = sys.argv[1]
    print(path_src)

    if os.path.isdir(path_src):
        files = os.listdir(path_src)
        for i,filename in enumerate(files):
            print("process {}, {} / {}".format(filename, i+1, len(files)))
            process(os.path.join(path_src, filename))
    else:
        process(path_src)
    
    print("done!")

