# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import requests
from requests import Response
import os
import sys


def process(path_src):

    workbook_src = load_workbook(path_src, data_only=True)
    for shert_name in workbook_src.sheetnames:
        sheet_src = workbook_src[shert_name]
        for col in range(1, sheet_src.max_column + 1):
            letter = get_column_letter(col)
            cell_name = letter + "1" # 取第一列
            cell_value = str(sheet_src[cell_name].value)
            if "TypIDVal" in cell_value:
                print(f"{path_src} {cell_value}")

    # workbook_src.save(path_src) # 储存

if __name__ == "__main__":
    path_src = sys.argv[1]
    print(path_src)

    if os.path.isdir(path_src):
        files = os.listdir(path_src)
        for i,filename in enumerate(files):
            print("process {}, {} / {}".format(filename, i+1, len(files)))
            process(os.path.join(path_src, filename))
    else:
        process(path_src)
    
    print("done!")

