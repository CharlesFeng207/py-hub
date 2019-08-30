# coding:utf-8
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import requests
from requests import Response
import os
import sys

       
if __name__ == "__main__":
    path_src = "Language.xlsx"
    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    target_letter = "B"
    start_row = 3

    pendingList = []
    for row in range(start_row, sheet_src.max_row + 1):
        cell = sheet_src["{}{}".format(target_letter, row)]
        if cell.value is None:
            continue

        pendingList.append("{}".format(str(cell.value)))
    
    print("count:{}".format(len(pendingList)))
    with open("output.txt", "w+") as f:
        f.write("\n@\n".join(pendingList))
    print("done!")

