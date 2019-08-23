from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import requests
from requests import Response
import os
import sys


def process(path_src):
    name_src = os.path.basename(path_src)

    workbook_src = load_workbook(path_src)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    target_cols = []
    i = 1
    while True:
        cell = sheet_src["{}{}".format(get_column_letter(i), 2)]
        if cell.value is None:
            break
        if cell.comment is not None and "@language" in cell.comment.content:
            target_cols.append(get_column_letter(i))
        i+=1

    print(target_cols)

    for target_letter in target_cols:
        name_cell = sheet_src["{}{}".format(target_letter, 1)]
        name_cell.value = name_cell.value + "_id"
        
        type_cell = sheet_src["{}{}".format(target_letter, 2)]
        type_cell.comment = Comment("language done!", "excel_tool")
        type_cell.value = "int32"

        for row in range(5, sheet_src.max_row + 1):
            cell = sheet_src["{}{}".format(target_letter, row)]
            print("{} cur:{} total:{} value:{}".format(name_src, row, sheet_src.max_row, cell.value))
            if cell.value is not None:
                respond = None
                while True:
                    isLast = row == sheet_src.max_row
                    respond = requests.get('http://139.155.88.114:5000/query', {"content":cell.value, "save":isLast})
                    if respond.status_code == 200:
                        break
                cell.comment = Comment(cell.value, "excel_tool")
                cell.value = respond.text
            else:
                cell.value = 0
    workbook_src.save(name_src)

if __name__ == "__main__":
    path_src = sys.argv[1]
    print(path_src)

    if os.path.isdir(path_src):
        files = os.listdir(path_src)
        for i,filename in enumerate(files):
            print("process {}, {} / {}".format(filename, i+1, len(files)))
            process(os.path.join(path_src, filename))
    else:
        process(path_src)
    
    print("done!")

