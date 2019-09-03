# -*- coding: utf-8 -*-
import sys
import uuid
import requests
import hashlib
import time
import json
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import os
from openpyxl.styles import PatternFill, colors


if __name__ == "__main__":

    path_src = "Language.xlsx"
    workbook_src = load_workbook(path_src, data_only=True)
    shert_name = list(filter(lambda x:x[0] != '@' and x[0] != '#', workbook_src.sheetnames))[0]
    print(shert_name)

    sheet_src = workbook_src[shert_name]

    target_letter = "B"
    start_row = 3
    jobs = 15

    total = 0
    for row in range(start_row, sheet_src.max_row + 1):
        cell = sheet_src["{}{}".format(target_letter, row)]
        if cell.value is None:
            continue

        total += len(str(cell.value))
        # if cell.fill.end_color.index != colors.BLACK:
        #     print(row)
   
    eachJobCount = total / jobs
    print("total:{} count:{} each:{}".format(total, jobs, eachJobCount))
    
    pending = []
    curNum = 0
    index = 1
    for row in range(start_row, sheet_src.max_row + 1):
        cell = sheet_src["{}{}".format(target_letter, row)]
        if cell.value is None:
            continue
        
        pending.append(row)
        curNum += len(str(cell.value))
        print("curNum: {} index:{}".format(curNum, index))
        
        if curNum > eachJobCount or row == sheet_src.max_row:
            print("make {}".format(index))
            nwb = Workbook()
            nwb.create_sheet("work", 0)
            nsheet=nwb["work"]

            for i, r in enumerate(pending):
                nsheet["A{}".format(i+1)].value = sheet_src["A{}".format(r)].value #id
                nsheet["B{}".format(i+1)].value = sheet_src["B{}".format(r)].value #cn
                nsheet["C{}".format(i+1)].value = sheet_src["C{}".format(r)].value #en

            if not os.path.exists("jobs"):
                os.mkdir("jobs")
            
            nwb.save(os.path.join("jobs", "{}.xlsx".format(index)))
            index += 1
            pending = []
            curNum = 0
        
    print("done!")